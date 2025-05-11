#!/usr/bin/env python3
"""
Convert Timetable CSV to Excel

This script converts the master_timetable.csv file to a properly formatted Excel file.
"""

import pandas as pd
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def convert_to_excel():
    """Convert the master timetable CSV to a formatted Excel file"""
    # Check if master timetable exists
    if not os.path.exists('master_timetable.csv'):
        print("Error: master_timetable.csv not found.")
        print("Please run the timetable generator first.")
        return
    
    print("Converting master_timetable.csv to Excel format...")
    
    # Load the CSV data
    df = pd.read_csv('master_timetable.csv')
    
    # Create an Excel writer
    output_file = 'master_timetable.xlsx'
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Write the data to Excel
    df.to_excel(writer, sheet_name='Master Timetable', index=False)
    
    # Get the workbook and the worksheet
    workbook = writer.book
    worksheet = writer.sheets['Master Timetable']
    
    # Define styles
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Set column widths
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    widths = [15, 8, 15, 20, 40, 20, 15]
    
    for i, width in enumerate(widths):
        if i < len(columns):
            col = columns[i]
            worksheet.column_dimensions[col].width = width
    
    # Apply styles to header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = cell_alignment
        cell.border = border
    
    # Apply styles to data cells
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = cell_alignment
            cell.border = border
    
    # Create additional sheet with pivot table view
    pivot_df = df.pivot_table(
        index=['Day', 'Slot', 'Time'],
        columns=['Course', 'Teacher', 'Room'],
        aggfunc='size',
        fill_value=0
    ).reset_index()
    
    # Save the workbook
    writer.close()
    
    print(f"Excel file created: {output_file}")

if __name__ == "__main__":
    convert_to_excel() 